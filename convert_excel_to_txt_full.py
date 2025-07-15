from openpyxl import load_workbook

input_excel = "file.xlsx"     # <-- your 6GB Excel file
output_txt = "pointcloud_input.txt"

# Load workbook and worksheet (first sheet assumed)
wb = load_workbook(filename=input_excel, read_only=True)
ws = wb.active  # use wb[sheetname] if you know it

with open(output_txt, "w") as out:
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        x, y, z = row
        out.write(f"{x} {y} {z}\n")

print("✅ Full dataset exported to pointcloud_input.txt without opening Excel")
